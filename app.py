import pandas as pd
import numpy as np
import streamlit as st
import plotly.express as px
import openpyxl
from io import BytesIO
import requests
from fpdf import FPDF
from datetime import datetime

st.set_page_config(page_title="Análise de Serviços Técnicos", layout="wide")

FORMAS_PAGAMENTO_VALIDAS = [
    'Check', 'American Express', 'Apple Pay', 'Discover',
    'Master Card', 'Visa', 'Zelle', 'Cash', 'Invoice'
]

INVALID_CLIENTS = ['SERVICES IN:', 'BNS PROFIT:', 'Total']


def format_currency(value):
    """Formata valores como moeda USD com 2 casas decimais"""
    if pd.isna(value):
        return None
    return f"${value:,.2f}"


def create_pdf(data):
    """Cria um PDF com os dados da página principal"""
    pdf = FPDF(orientation='P', unit='mm', format='A4')
    pdf.add_page()
    pdf.set_font("Arial", size=10)

    # Configurações de margem
    left_margin = 10
    right_margin = 10
    pdf.set_left_margin(left_margin)
    pdf.set_right_margin(right_margin)
    page_width = pdf.w - left_margin - right_margin

    # Adiciona título
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(page_width, 10, txt="BNS - PORTAL DE ANÁLISES DE DADOS FINANCEIROS", ln=1, align='C')
    pdf.ln(5)

    # Adiciona data de geração
    pdf.set_font("Arial", size=10)
    pdf.cell(page_width, 10, txt=f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", ln=1, align='R')
    pdf.ln(10)

    # Seção 1: Métricas Gerais
    pdf.set_font("Arial", 'B', 14)
    pdf.cell(page_width, 10, txt="1. Métricas Gerais", ln=1)
    pdf.set_font("Arial", size=10)

    completed_services = data[data['Realizado']]
    not_completed = data[(data['Realizado'] == False) & (data['Cliente'].notna())]

    metrics = [
        ("Atendimentos Realizados", len(completed_services)),
        ("Atendimentos Não Realizados", len(not_completed)),
        ("Total em Serviços", format_currency(completed_services['Serviço'].sum())),
        ("Total em Gorjetas", format_currency(completed_services['Gorjeta'].sum())),
        ("Lucro da Empresa", format_currency(completed_services['Lucro Empresa'].sum()))
    ]

    for metric, value in metrics:
        pdf.cell(page_width / 2, 10, txt=f"{metric}:", ln=0)
        pdf.cell(page_width / 2, 10, txt=str(value), ln=1)

    pdf.ln(10)

    # Seção 2: Resumo por Técnico
    pdf.set_font("Arial", 'B', 14)
    pdf.cell(page_width, 10, txt="2. Resumo por Técnico", ln=1)

    # Prepara dados para a tabela
    tech_summary = completed_services.groupby(['Nome', 'Categoria']).agg({
        'Serviço': 'sum',
        'Gorjeta': 'sum',
        'Pagamento Tecnico': 'sum',
        'Lucro Empresa': 'sum',
        'Cliente': 'count'
    }).reset_index()

    tech_summary.columns = ['Técnico', 'Categoria', 'Total Serviços', 'Total Gorjetas',
                            'Total Pagamento', 'Lucro Empresa', 'Atendimentos']

    # Adiciona tabela de técnicos
    pdf.set_font("Arial", size=8)  # Fonte menor para caber mais colunas
    col_widths = [30, 25, 25, 25, 25, 25]  # Larguras ajustadas para caber na página

    # Cabeçalho da tabela
    headers = ["Técnico", "Categoria", "Serviços", "Gorjetas", "Pagamento", "Lucro"]
    for i, header in enumerate(headers):
        pdf.cell(col_widths[i], 10, txt=header, border=1, align='C')
    pdf.ln()

    # Linhas da tabela
    for _, row in tech_summary.iterrows():
        # Quebra de linha se o nome for muito longo
        tech_name = str(row['Técnico'])[:15] + '...' if len(str(row['Técnico'])) > 15 else str(row['Técnico'])
        pdf.cell(col_widths[0], 10, txt=tech_name, border=1)
        pdf.cell(col_widths[1], 10, txt=str(row['Categoria'])[:10], border=1)  # Limita categoria
        pdf.cell(col_widths[2], 10, txt=format_currency(row['Total Serviços']), border=1, align='R')
        pdf.cell(col_widths[3], 10, txt=format_currency(row['Total Gorjetas']), border=1, align='R')
        pdf.cell(col_widths[4], 10, txt=format_currency(row['Total Pagamento']), border=1, align='R')
        pdf.cell(col_widths[5], 10, txt=format_currency(row['Lucro Empresa']), border=1, align='R')
        pdf.ln()

    pdf.ln(10)

    # Seção 3: Métodos de Pagamento
    pdf.set_font("Arial", 'B', 14)
    pdf.cell(page_width, 10, txt="3. Métodos de Pagamento", ln=1)
    pdf.set_font("Arial", size=10)

    valid_payments = completed_services[completed_services['Pagamento'].isin(FORMAS_PAGAMENTO_VALIDAS)]

    if not valid_payments.empty:
        payment_methods = valid_payments.groupby('Pagamento').agg({
            'Serviço': ['sum', 'count'],
            'Gorjeta': 'sum',
            'Lucro Empresa': 'sum'
        }).reset_index()

        payment_methods.columns = ['Método', 'Total Serviços', 'Qtd Usos', 'Total Gorjetas', 'Lucro Empresa']
        payment_methods['Total Geral'] = payment_methods['Total Serviços'] + payment_methods['Total Gorjetas']

        # Tabela de métodos de pagamento
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(page_width, 10, txt="Resumo por Método de Pagamento:", ln=1)
        pdf.set_font("Arial", size=8)  # Fonte menor para tabela

        # Cabeçalho
        headers = ["Método", "Usos", "Serviços", "Gorjetas", "Total", "Lucro"]
        col_widths_payments = [30, 20, 25, 25, 25, 25]  # Larguras ajustadas

        for i, header in enumerate(headers):
            pdf.cell(col_widths_payments[i], 10, txt=header, border=1, align='C')
        pdf.ln()

        # Linhas
        for _, row in payment_methods.iterrows():
            pdf.cell(col_widths_payments[0], 10, txt=str(row['Método'])[:12], border=1)  # Limita método
            pdf.cell(col_widths_payments[1], 10, txt=str(row['Qtd Usos']), border=1, align='C')
            pdf.cell(col_widths_payments[2], 10, txt=format_currency(row['Total Serviços']), border=1, align='R')
            pdf.cell(col_widths_payments[3], 10, txt=format_currency(row['Total Gorjetas']), border=1, align='R')
            pdf.cell(col_widths_payments[4], 10, txt=format_currency(row['Total Geral']), border=1, align='R')
            pdf.cell(col_widths_payments[5], 10, txt=format_currency(row['Lucro Empresa']), border=1, align='R')
            pdf.ln()

        # Adiciona porcentagem de uso
        pdf.ln(5)
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(page_width, 10, txt="Distribuição por Método de Pagamento:", ln=1)
        pdf.set_font("Arial", size=10)

        total_usos = payment_methods['Qtd Usos'].sum()
        for _, row in payment_methods.iterrows():
            percent = (row['Qtd Usos'] / total_usos * 100)
            pdf.cell(page_width / 2, 10, txt=f"{row['Método']}:", ln=0)
            pdf.cell(page_width / 2, 10, txt=f"{percent:.1f}% ({row['Qtd Usos']} usos)", ln=1)

    pdf.ln(10)

    # Seção 4: Atendimentos por Dia da Semana
    pdf.set_font("Arial", 'B', 14)
    pdf.cell(page_width, 10, txt="4. Atendimentos por Dia da Semana", ln=1)
    pdf.set_font("Arial", size=10)

    day_summary = completed_services.groupby('Dia').agg({
        'Serviço': ['count', 'sum'],
        'Gorjeta': 'sum',
        'Lucro Empresa': 'sum'
    }).reset_index()

    day_summary.columns = ['Dia', 'Atendimentos', 'Total Serviços', 'Total Gorjetas', 'Lucro Empresa']

    # Ordena os dias corretamente
    day_order = ['Domingo', 'Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado']
    day_summary['Dia'] = pd.Categorical(day_summary['Dia'], categories=day_order, ordered=True)
    day_summary = day_summary.sort_values('Dia')

    # Tabela de dias
    col_widths_days = [30, 25, 30, 30, 30]  # Larguras ajustadas
    headers = ["Dia", "Atend.", "Serviços", "Gorjetas", "Lucro"]

    pdf.set_font("Arial", size=8)
    for i, header in enumerate(headers):
        pdf.cell(col_widths_days[i], 10, txt=header, border=1, align='C')
    pdf.ln()

    for _, row in day_summary.iterrows():
        pdf.cell(col_widths_days[0], 10, txt=str(row['Dia']), border=1)
        pdf.cell(col_widths_days[1], 10, txt=str(row['Atendimentos']), border=1, align='C')
        pdf.cell(col_widths_days[2], 10, txt=format_currency(row['Total Serviços']), border=1, align='R')
        pdf.cell(col_widths_days[3], 10, txt=format_currency(row['Total Gorjetas']), border=1, align='R')
        pdf.cell(col_widths_days[4], 10, txt=format_currency(row['Lucro Empresa']), border=1, align='R')
        pdf.ln()

    pdf.ln(10)

    # Seção 5: Atendimentos Não Realizados
    if len(not_completed) > 0:
        pdf.set_font("Arial", 'B', 14)
        pdf.cell(page_width, 10, txt="5. Atendimentos Não Realizados", ln=1)
        pdf.set_font("Arial", size=10)

        pdf.cell(page_width, 10, txt=f"Total de atendimentos não realizados: {len(not_completed)}", ln=1)

        # Lista os primeiros 10 atendimentos não realizados
        pdf.set_font("Arial", size=8)
        for idx, row in not_completed.head(10).iterrows():
            pdf.cell(page_width, 10,
                     txt=f"- {row['Nome']} | {row['Dia']} {row['Data'].strftime('%d/%m')} | {row['Cliente']}",
                     ln=1)

    return pdf


def create_tech_payment_receipt(tech_data, tech_name, week):
    """Cria um PDF com o recibo de pagamento detalhado para o técnico com papel timbrado"""
    pdf = FPDF(orientation='P', unit='mm', format='A4')
    pdf.add_page()

    # Configurações de margem
    left_margin = 15
    right_margin = 15
    pdf.set_left_margin(left_margin)
    pdf.set_right_margin(right_margin)
    page_width = pdf.w - left_margin - right_margin
    page_height = pdf.h

    # Calcular intervalo de datas
    min_date = tech_data['Data'].min().strftime('%m/%d/%y')
    max_date = tech_data['Data'].max().strftime('%m/%d/%y')
    date_range = f"{min_date} to {max_date}"

    # Restante do conteúdo do recibo
    pdf.set_font("Arial", 'B', 18)
    pdf.cell(page_width, 10, txt="TECHNICIAN PAYMENT RECEIPT", ln=1, align='C')
    pdf.ln(9)

    # Informações do técnico e semana
    pdf.set_font("Arial", size=10)
    pdf.cell(page_width, 8, txt=f"Technician: {tech_name}", ln=1)
    pdf.cell(page_width, 8, txt=f"Reference: {date_range}", ln=1)  # Alterado para mostrar intervalo de datas
    pdf.cell(page_width, 8, txt=f"Date of issue: {datetime.now().strftime('%m/%d/%Y')}", ln=1)
    pdf.ln(10)

    # Resumo de atendimentos
    pdf.set_font("Arial", 'B', 14)
    pdf.cell(page_width, 10, txt="SUMMARY OF SERVICES", ln=1)
    pdf.set_font("Arial", size=10)

    total_services = tech_data['Serviço'].sum()
    total_tips = tech_data['Gorjeta'].sum()
    total_payment = tech_data['Pagamento Tecnico'].sum()

    # Formata os valores
    def format_value(value):
        return f"${value:,.2f}" if isinstance(value, (int, float)) else str(value)

    # Tabela de resumo
    col_widths = [page_width / 2, page_width / 2]

    pdf.cell(col_widths[0], 10, txt="Total Schedules:", border='B', ln=0)
    pdf.cell(col_widths[1], 10, txt=str(len(tech_data)), border='B', ln=1, align='R')

    pdf.cell(col_widths[0], 10, txt="Total in Services:", border='B', ln=0)
    pdf.cell(col_widths[1], 10, txt=format_value(total_services), border='B', ln=1, align='R')

    pdf.cell(col_widths[0], 10, txt="Total in Tips:", border='B', ln=0)
    pdf.cell(col_widths[1], 10, txt=format_value(total_tips), border='B', ln=1, align='R')

    pdf.set_font("Arial", 'B', 12)
    pdf.cell(col_widths[0], 10, txt="Total Payment", border='B', ln=0)
    pdf.cell(col_widths[1], 10, txt=format_value(total_payment), border='B', ln=1, align='R')
    pdf.set_font("Arial", size=10)

    pdf.ln(15)

    # Detalhes por dia
    pdf.set_font("Arial", 'B', 14)
    pdf.cell(page_width, 10, txt="DETAILS BY DAY", ln=1)
    pdf.set_font("Arial", size=8)

    # Agrupar por dia
    day_details = tech_data.groupby('Dia').agg({
        'Serviço': 'sum',
        'Gorjeta': 'sum',
        'Cliente': 'count',
        'Pagamento': lambda x: ', '.join([str(p) for p in x.unique() if pd.notna(p)])
    }).reset_index()

    # Mapear os dias para inglês
    day_mapping = {
        'Domingo': 'Sun',
        'Segunda': 'Mon',
        'Terça': 'Tue',
        'Quarta': 'Wed',
        'Quinta': 'Thu',
        'Sexta': 'Fri',
        'Sábado': 'Sat'
    }
    day_details['Dia'] = day_details['Dia'].map(day_mapping)

    # Ordenar os dias corretamente
    day_order = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
    day_details['Dia'] = pd.Categorical(day_details['Dia'], categories=day_order, ordered=True)
    day_details = day_details.sort_values('Dia')

    # Cabeçalho da tabela
    col_widths = [46, 46, 47, 46]  # Ajustado para caber no timbrado
    headers = ["Day", "Showed", "Services", "Tips"]

    pdf.set_font("Arial", 'B', 7)
    for i, header in enumerate(headers):
        pdf.cell(col_widths[i], 6, txt=header, border=1, align='C')
    pdf.ln()

    # Linhas da tabela
    pdf.set_font("Arial", size=7)
    for _, row in day_details.iterrows():
        pdf.cell(col_widths[0], 6, txt=str(row['Dia']), border=1)  # Dia em inglês
        pdf.cell(col_widths[1], 6, txt=str(row['Cliente']), border=1, align='C')
        pdf.cell(col_widths[2], 6, txt=format_value(row['Serviço']), border=1, align='R')
        pdf.cell(col_widths[3], 6, txt=format_value(row['Gorjeta']), border=1, align='R')

        pdf.ln()

    pdf.ln(10)

    # Detalhes dos atendimentos (se couber na página)
    if pdf.get_y() < page_height - 50:  # Verifica se há espaço na página
        pdf.set_font("Arial", 'B', 14)
        pdf.cell(page_width, 10, txt="SERVICE DETAILS", ln=1)
        pdf.set_font("Arial", size=7)

        # Ordenar por data e dia
        tech_data_sorted = tech_data.sort_values(['Data', 'Dia'])

        # Cabeçalho da tabela detalhada
        col_widths_detailed = [25, 25, 60, 25, 25, 25]  # Ajustado para timbrado
        headers_detailed = ["Date", "Day", "Customer", "Service", "Tips", "Payment"]

        pdf.set_font("Arial", 'B', 7)
        for i, header in enumerate(headers_detailed):
            pdf.cell(col_widths_detailed[i], 6, txt=header, border=1, align='C')
        pdf.ln()

        # Linhas da tabela detalhada
        pdf.set_font("Arial", size=7)
        for _, row in tech_data_sorted.iterrows():
            if pdf.get_y() > page_height - 20:  # Verifica fim da página
                pdf.add_page()
                # Adiciona timbrado na nova página
                try:
                    pdf.image(timbrado_url, x=0, y=0, w=pdf.w, h=pdf.h, type='PNG')
                    pdf.set_fill_color(255, 255, 255, 80)
                    pdf.rect(0, 0, pdf.w, pdf.h, 'F')
                except:
                    pass
                pdf.set_y(30)

                # Recria cabeçalho da tabela
                pdf.set_font("Arial", 'B', 7)
                for i, header in enumerate(headers_detailed):
                    pdf.cell(col_widths_detailed[i], 8, txt=header, border=1, align='C')
                pdf.ln()
                pdf.set_font("Arial", size=6)

            # Data
            pdf.cell(col_widths_detailed[0], 6, txt=row['Data'].strftime('%d/%m'), border=1)
            # Dia (convertido para inglês)
            day_english = day_mapping.get(row['Dia'], row['Dia'])
            pdf.cell(col_widths_detailed[1], 6, txt=day_english, border=1)
            # Cliente
            client_name = str(row['Cliente'])[:20] + '...' if len(str(row['Cliente'])) > 20 else str(row['Cliente'])
            pdf.cell(col_widths_detailed[2], 6, txt=client_name, border=1)
            # Serviço
            pdf.cell(col_widths_detailed[3], 6, txt=format_value(row['Serviço']), border=1, align='R')
            # Gorjeta
            pdf.cell(col_widths_detailed[4], 6, txt=format_value(row['Gorjeta']), border=1, align='R')
            # Pagamento
            payment = str(row['Pagamento']) if pd.notna(row['Pagamento']) else "-"
            pdf.cell(col_widths_detailed[5], 6, txt=payment[:12], border=1)
            pdf.ln()

    # Informação da empresa no rodapé
    pdf.set_font("Arial", size=8)
    pdf.cell(page_width, 5, txt="BRIGHT N SHINE PET DENTAL LLC", ln=1, align='C')
    pdf.cell(page_width, 5, txt="(407)259-7897", ln=1, align='C')

    return pdf


def process_spreadsheet(file):
    all_weeks_data = {}
    if isinstance(file, str) and file.startswith('http'):
        response = requests.get(file)
        file = BytesIO(response.content)
    elif isinstance(file, BytesIO):
        file.seek(0)

    xls = pd.ExcelFile(file)
    for sheet_name in xls.sheet_names:
        if sheet_name.startswith('WEEK'):
            df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
            technician_blocks = []
            current_block = []
            collecting = False
            for idx, row in df.iterrows():
                if any('NAME:' in str(cell) for cell in row.values):
                    if current_block:
                        technician_blocks.append(current_block)
                        current_block = []
                    collecting = True
                if collecting:
                    current_block.append(row)
            if current_block:
                technician_blocks.append(current_block)

            week_data = []
            for block in technician_blocks:
                name_row = next((row for row in block if any('NAME:' in str(cell) for cell in row.values)), None)

                if name_row is None:
                    continue
                name_col = next(
                    (i for i, cell in enumerate(name_row.values) if isinstance(cell, str) and 'NAME:' in cell), None)

                technician_info = {
                    'Semana': sheet_name,
                    'Nome': name_row[name_col + 1] if name_col is not None else None,
                    'Categoria': name_row[name_col + 3] if name_col is not None else None,
                    'Origem': name_row[name_col + 5] if name_col is not None and 'From:' in str(
                        name_row[name_col + 4]) else None
                }

                header_row = next((i for i, row in enumerate(block) if all(
                    keyword in str(row.values) for keyword in ['Schedule', 'DATE', 'SERVICE'])), None)
                if header_row is None:
                    continue

                days_data = []
                for i in range(header_row + 1, len(block)):
                    day_row = block[i]
                    for day_idx, day_col in enumerate(
                            [(1, 9), (10, 18), (19, 27), (28, 36), (37, 45), (46, 54), (55, 63)]):
                        start_col, end_col = day_col
                        day_data = day_row[start_col:end_col + 1].values
                        client_name = str(day_data[0]).strip() if pd.notna(day_data[0]) else ''

                        # Ignorar linhas inválidas
                        if not client_name or client_name.upper() in [c.upper() for c in INVALID_CLIENTS]:
                            continue

                        # Verificar se é um atendimento válido (tem valor de serviço)
                        if pd.notna(day_data[2]) and str(day_data[2]).strip() and str(day_data[2]).strip() != 'nan':
                            try:
                                service_value = float(day_data[2])
                            except:
                                service_value = np.nan
                            if not np.isnan(service_value):
                                pagamento = day_data[5] if pd.notna(day_data[5]) and str(
                                    day_data[5]).strip() in FORMAS_PAGAMENTO_VALIDAS else None
                                tip_value = float(day_data[3]) if pd.notna(day_data[3]) else 0
                                day_info = {
                                    'Dia': ['Domingo', 'Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado'][
                                        day_idx],
                                    'Data': day_data[1],
                                    'Cliente': client_name,
                                    'Serviço': service_value,
                                    'Gorjeta': tip_value,
                                    'Pets': day_data[4] if pd.notna(day_data[4]) else 0,
                                    'Pagamento': pagamento,
                                    'ID Pagamento': day_data[6] if pd.notna(day_data[6]) else None,
                                    'Verificado': day_data[7] if pd.notna(day_data[7]) else False,
                                    'Realizado': True
                                }
                                days_data.append({**technician_info, **day_info})
                        elif pd.notna(day_data[0]):
                            if client_name.upper() in [c.upper() for c in INVALID_CLIENTS]:
                                continue
                            day_info = {
                                'Dia': ['Domingo', 'Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado'][day_idx],
                                'Data': day_data[1],
                                'Cliente': client_name,
                                'Serviço': 0,
                                'Gorjeta': 0,
                                'Pets': 0,
                                'Pagamento': None,
                                'ID Pagamento': None,
                                'Verificado': False,
                                'Realizado': False
                            }
                            days_data.append({**technician_info, **day_info})

                week_data.extend(days_data)
            if week_data:
                all_weeks_data[sheet_name] = pd.DataFrame(week_data)

    if all_weeks_data:
        combined_data = pd.concat(all_weeks_data.values(), ignore_index=True)
        combined_data['Data'] = pd.to_datetime(combined_data['Data'], errors='coerce')
        combined_data['Serviço'] = pd.to_numeric(combined_data['Serviço'], errors='coerce')
        combined_data['Gorjeta'] = pd.to_numeric(combined_data['Gorjeta'], errors='coerce').fillna(0)
        combined_data['Pets'] = pd.to_numeric(combined_data['Pets'], errors='coerce').fillna(0)
        combined_data = combined_data.dropna(subset=['Data'])
        combined_data = combined_data[
            ~combined_data['Cliente'].astype(str).str.strip().str.upper().isin([c.upper() for c in INVALID_CLIENTS])]
        return combined_data
    return pd.DataFrame()


def calcular_pagamento_individual(row, weekly_data):
    """Calcula o pagamento individual de cada atendimento"""
    tech_week_data = weekly_data[
        (weekly_data['Nome'] == row['Nome']) &
        (weekly_data['Semana'] == row['Semana'])
        ]

    if len(tech_week_data) == 0:
        return pd.Series([0, row['Serviço'] + row['Gorjeta']])

    total_pagamento = tech_week_data['Pagamento Tecnico'].iloc[
        0] if 'Pagamento Tecnico' in tech_week_data.columns else 0
    total_servico = tech_week_data['Serviço'].sum()

    if total_servico == 0:
        return pd.Series([0, row['Serviço'] + row['Gorjeta']])

    # Pagamento proporcional ao serviço realizado neste atendimento
    try:
        pagamento = (row['Serviço'] / total_servico) * total_pagamento
        lucro = row['Serviço'] + row['Gorjeta'] - pagamento
    except:
        pagamento = 0
        lucro = row['Serviço'] + row['Gorjeta']

    return pd.Series([pagamento, lucro])


def calcular_pagamento_semanal(row):
    """Calcula o pagamento semanal baseado na categoria"""
    categoria = row['Categoria']
    servico = row['Serviço']
    gorjeta = row['Gorjeta']
    dias_trabalhados = row['Dias Trabalhados']

    if categoria == 'Registering':
        pagamento = 0.00
        lucro = servico + gorjeta
    elif categoria == 'Technician':
        pagamento = servico * 0.20 + gorjeta
        lucro = servico * 0.80
    elif categoria == 'Training':
        pagamento = 80 * dias_trabalhados  # $80 por dia trabalhado
        lucro = servico + gorjeta - pagamento
    elif categoria == 'Coordinator':
        pagamento = servico * 0.25 + gorjeta
        lucro = servico * 0.75
    elif categoria == 'Started':
        # Aplicar cálculo por semana individualmente
        valor_comissao = servico * 0.20 + gorjeta
        valor_minimo = 150 * dias_trabalhados
        pagamento = max(valor_minimo, valor_comissao)
        lucro = servico + gorjeta - pagamento
    else:
        pagamento = 0
        lucro = servico + gorjeta

    return pd.Series([pagamento, lucro])


# Configuração da sidebar
st.sidebar.markdown("""
<div style="text-align: center; margin-bottom: 20px;">
    <img src="https://i.imgur.com/tlb2Bcy.png" 
         alt="Logo da Empresa" 
</div>
""", unsafe_allow_html=True)

st.sidebar.title("🔍 Filtros")

# Main content
st.title("📊 BNS - PORTAL DE ANÁLISES DE DADOS FINANCEIROS")

uploaded_files = st.sidebar.file_uploader("Carregue uma ou mais planilhas Excel", type=['xlsx'],
                                          accept_multiple_files=True)
url_input = st.sidebar.text_input("Ou cole a URL de uma planilha online")

all_dataframes = []
if uploaded_files or url_input:
    files_to_process = uploaded_files if uploaded_files else [url_input]
    for file in files_to_process:
        df = process_spreadsheet(file)
        if not df.empty:
            all_dataframes.append(df)

    if all_dataframes:
        data = pd.concat(all_dataframes, ignore_index=True)
        data = data[data['Nome'].notna() & (data['Nome'].astype(str).str.strip() != '')]
        data = data[~data['Cliente'].astype(str).str.strip().str.upper().isin([c.upper() for c in INVALID_CLIENTS])]

        # Definir opções de filtro
        weeks = data['Semana'].unique()
        technicians = data['Nome'].unique()
        categories = data['Categoria'].unique()

        # Filtros na sidebar
        st.sidebar.header("Filtrar por:")

        # Filtrar por abas (Semana)
        selected_weeks = st.sidebar.multiselect(
            "Selecione as abas (semanas):",
            options=weeks,
            default=list(weeks)
        )

        # Filtrar por técnico
        selected_techs = st.sidebar.multiselect(
            "Selecione os técnicos:",
            options=technicians,
            default=list(technicians)
        )

        # Filtrar por categoria
        selected_categories = st.sidebar.multiselect(
            "Selecione as categorias:",
            options=categories,
            default=list(categories))

        # Aplicar filtros
        if selected_weeks:
            data = data[data['Semana'].isin(selected_weeks)]
        if selected_techs:
            data = data[data['Nome'].isin(selected_techs)]
        if selected_categories:
            data = data[data['Categoria'].isin(selected_categories)]

        if data.empty:
            st.warning("Nenhum dado encontrado com os filtros selecionados.")
            st.stop()

        st.success("✅ Planilhas processadas com sucesso!")

        if st.checkbox("🔍 Mostrar dados brutos"):
            st.dataframe(data)

        st.header("📈 Métricas Gerais")
        completed_services = data[data['Realizado']]
        not_completed = data[(data['Realizado'] == False) & (data['Cliente'].notna())]

        # Calcular dias trabalhados corretamente (1 por dia com atendimento, por técnico por semana)
        dias_trabalhados = completed_services.groupby(['Nome', 'Semana', 'Data']).size().reset_index()
        dias_trabalhados = dias_trabalhados.groupby(['Nome', 'Semana']).size().reset_index(name='Dias Trabalhados')

        # Agrupar por técnico e semana para calcular totais
        weekly_totals = completed_services.groupby(['Nome', 'Semana', 'Categoria']).agg({
            'Serviço': 'sum',
            'Gorjeta': 'sum',
            'Dia': 'count'
        }).reset_index()

        # Juntar com os dias trabalhados corretamente calculados
        weekly_totals = pd.merge(weekly_totals, dias_trabalhados, on=['Nome', 'Semana'], how='left')

        # Aplicar cálculo de pagamento e lucro semanal
        weekly_totals[['Pagamento Tecnico', 'Lucro Empresa']] = weekly_totals.apply(
            calcular_pagamento_semanal, axis=1)

        # Aplicar cálculo proporcional para cada atendimento
        if 'Pagamento Tecnico' not in completed_services.columns:
            completed_services[['Pagamento Tecnico', 'Lucro Empresa']] = completed_services.apply(
                lambda x: calcular_pagamento_individual(x, weekly_totals), axis=1)

        total_lucro = completed_services['Lucro Empresa'].sum() if 'Lucro Empresa' in completed_services.columns else 0
        total_pagamentos = completed_services[
            'Pagamento Tecnico'].sum() if 'Pagamento Tecnico' in completed_services.columns else 0

        col1, col2, col3, col4, col5 = st.columns(5)
        col1.metric("Realizados", len(completed_services))
        col2.metric("Não Realizados", len(not_completed))
        col3.metric("Total em Serviços", format_currency(completed_services['Serviço'].sum()))
        col4.metric("Total em Gorjetas", format_currency(completed_services['Gorjeta'].sum()))
        col5.metric("Lucro da Empresa", format_currency(total_lucro))

        # LAYOUT COM COLUNAS
        col_calculos, col_analise = st.columns([1, 2])

        with col_calculos:
            st.header("Cálculos Semanais")

            # Formatar valores monetários para exibição
            weekly_totals_display = weekly_totals.copy()
            weekly_totals_display['Serviço'] = weekly_totals_display['Serviço'].apply(format_currency)
            weekly_totals_display['Gorjeta'] = weekly_totals_display['Gorjeta'].apply(format_currency)
            weekly_totals_display['Pagamento Tecnico'] = weekly_totals_display['Pagamento Tecnico'].apply(
                format_currency)
            weekly_totals_display['Lucro Empresa'] = weekly_totals_display['Lucro Empresa'].apply(format_currency)

            weekly_totals_display = weekly_totals_display.rename(columns={
                'Nome': 'Técnico',
                'Semana': 'Semana',
                'Categoria': 'Categoria',
                'Serviço': 'Total Serviços',
                'Gorjeta': 'Total Gorjetas',
                'Pagamento Tecnico': 'Pagamento Semanal',
                'Lucro Empresa': 'Lucro da Empresa',
                'Dias Trabalhados': 'Dias Trabalhados'
            })

            st.dataframe(weekly_totals_display)

        with col_analise:
            st.header("Análise por Técnico")

            # Agrupar por técnico e categoria
            tech_summary = weekly_totals.groupby(['Nome', 'Categoria']).agg({
                'Serviço': 'sum',
                'Gorjeta': 'sum',
                'Pagamento Tecnico': 'sum',
                'Lucro Empresa': 'sum',
                'Dia': 'sum',
                'Dias Trabalhados': 'sum'
            }).reset_index()

            # Ajustar nomes das colunas
            tech_summary.columns = ['Técnico', 'Categoria', 'Total Serviços',
                                    'Total Gorjetas', 'Total Pagamento', 'Lucro Empresa',
                                    'Atendimentos', 'Dias Trabalhados']

            tech_summary['Média Atendimento'] = tech_summary['Total Serviços'] / tech_summary['Atendimentos']
            tech_summary['Gorjeta Média'] = tech_summary['Total Gorjetas'] / tech_summary['Atendimentos']

            # Formatar valores monetários
            tech_summary['Total Serviços'] = tech_summary['Total Serviços'].apply(format_currency)
            tech_summary['Total Gorjetas'] = tech_summary['Total Gorjetas'].apply(format_currency)
            tech_summary['Total Pagamento'] = tech_summary['Total Pagamento'].apply(format_currency)
            tech_summary['Lucro Empresa'] = tech_summary['Lucro Empresa'].apply(format_currency)
            tech_summary['Média Atendimento'] = tech_summary['Média Atendimento'].apply(format_currency)
            tech_summary['Gorjeta Média'] = tech_summary['Gorjeta Média'].apply(format_currency)

            st.dataframe(tech_summary.sort_values('Atendimentos', ascending=False))

        st.subheader("📈 Evolução Semanal por Técnico")
        fig_evolucao = px.line(
            weekly_totals,
            x='Semana',
            y='Serviço',
            color='Nome',
            markers=True,
            title='Evolução de Serviços por Técnico',
            labels={'Serviço': 'Valor em Serviços ($)', 'Semana': 'Semana'}
        )
        fig_evolucao.update_traces(hovertemplate="<b>%{x}</b><br>Valor: $%{y:,.2f}")
        st.plotly_chart(fig_evolucao, use_container_width=True)

        # Técnico da Semana
        if len(selected_weeks) == 1:
            semana_atual = selected_weeks[0]
            tech_da_semana = \
                weekly_totals[weekly_totals['Semana'] == semana_atual].sort_values('Serviço',
                                                                                   ascending=False).iloc[0]

            st.subheader("🏆 Técnico da Semana")
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Técnico", tech_da_semana['Nome'])
            col2.metric("Total em Serviços", format_currency(tech_da_semana['Serviço']))
            col3.metric("Pagamento Semanal", format_currency(tech_da_semana['Pagamento Tecnico']))
            col4.metric("Lucro Empresa", format_currency(tech_da_semana['Lucro Empresa']))

        fig_pagamento = px.bar(
            weekly_totals,
            x='Pagamento Tecnico',
            y='Nome',
            color='Semana',
            barmode='group',
            title='Pagamento Semanal por Técnico',
            labels={'Pagamento Tecnico': 'Pagamento ($)', 'Nome': 'Técnico'}
        )
        fig_pagamento.update_traces(texttemplate='$%{x:,.2f}', textposition='outside')
        fig_pagamento.update_layout(hovermode="x unified")
        st.plotly_chart(fig_pagamento, use_container_width=True)

        # Gráfico de atendimentos por técnico
        tech_summary_graph = tech_summary.copy()
        tech_summary_graph['Atendimentos'] = pd.to_numeric(tech_summary_graph['Atendimentos'], errors='coerce')

        fig1 = px.bar(tech_summary_graph.sort_values('Atendimentos'),
                      x='Atendimentos', y='Técnico',
                      title='Atendimentos por Técnico',
                      color='Categoria',
                      labels={'Atendimentos': 'Quantidade'})
        fig1.update_traces(hovertemplate="<b>%{y}</b><br>Atendimentos: %{x}<br>Categoria: %{marker.color}")
        st.plotly_chart(fig1, use_container_width=True)

        # Gráfico de gorjetas por técnico
        fig2 = px.bar(tech_summary_graph.sort_values('Total Gorjetas'),
                      x='Total Gorjetas', y='Técnico',
                      title='Gorjetas por Técnico',
                      color='Categoria',
                      labels={'Total Gorjetas': 'Valor Gorjetas ($)'})
        fig2.update_traces(hovertemplate="<b>%{y}</b><br>Total Gorjetas: $%{x:,.2f}<br>Categoria: %{marker.color}")
        st.plotly_chart(fig2, use_container_width=True)

        st.header("⚠️ Atendimentos Não Realizados")
        if not not_completed.empty:
            st.warning(f"{len(not_completed)} atendimentos não realizados.")
            st.dataframe(not_completed[['Nome', 'Dia', 'Data', 'Cliente']])
        else:
            st.success("Todos os agendamentos foram realizados!")

        st.header("💳 Métodos de Pagamento")
        valid_payments = completed_services[completed_services['Pagamento'].isin(FORMAS_PAGAMENTO_VALIDAS)]
        invalid_payments = completed_services[
            ~completed_services['Pagamento'].isin(FORMAS_PAGAMENTO_VALIDAS) & completed_services['Pagamento'].notna()]

        # Criar colunas para métricas
        col1, col2, col3 = st.columns(3)
        col1.metric("Válidos", len(valid_payments))
        col2.metric("Inválidos", len(invalid_payments))
        col3.metric("Formas de Pagamento", len(valid_payments['Pagamento'].unique()))

        if not valid_payments.empty:
            # Criar dataframe com informações detalhadas
            payment_methods = valid_payments.groupby('Pagamento').agg({
                'Serviço': ['sum', 'count'],
                'Gorjeta': 'sum',
                'Cliente': 'count',
                'Lucro Empresa': 'sum'
            }).reset_index()

            # Renomear colunas para melhor visualização
            payment_methods.columns = ['Pagamento', 'Total Serviços', 'Qtd Usos', 'Total Gorjetas',
                                       'Total Atendimentos', 'Lucro Empresa']

            # Calcular valores totais
            payment_methods['Total Geral'] = payment_methods['Total Serviços'] + payment_methods['Total Gorjetas']

            # Calcular porcentagem de uso
            total_usos = payment_methods['Qtd Usos'].sum()
            payment_methods['% Uso'] = (payment_methods['Qtd Usos'] / total_usos * 100).round(2)

            # Formatar valores monetários
            payment_methods['Total Serviços'] = payment_methods['Total Serviços'].apply(format_currency)
            payment_methods['Total Gorjetas'] = payment_methods['Total Gorjetas'].apply(format_currency)
            payment_methods['Lucro Empresa'] = payment_methods['Lucro Empresa'].apply(format_currency)
            payment_methods['Total Geral'] = payment_methods['Total Geral'].apply(format_currency)
            payment_methods['% Uso'] = payment_methods['% Uso'].astype(str) + '%'

            # Mostrar tabela detalhada
            st.subheader("Detalhes por Método de Pagamento")
            st.dataframe(payment_methods.sort_values('Qtd Usos', ascending=False))

            # Criar gráficos
            tab1, tab2 = st.tabs(["Valor Total", "Quantidade de Usos"])

            with tab1:
                # Dataframe para gráfico (valores numéricos)
                payment_graph = valid_payments.groupby('Pagamento').agg({
                    'Serviço': 'sum',
                    'Gorjeta': 'sum',
                    'Lucro Empresa': 'sum'
                }).reset_index()
                payment_graph['Total'] = payment_graph['Serviço'] + payment_graph['Gorjeta']

                fig_total = px.bar(payment_graph.sort_values('Total'),
                                   x='Total', y='Pagamento',
                                   title='Valor Total por Método de Pagamento (Serviços + Gorjetas)',
                                   color='Serviço',
                                   color_continuous_scale='Peach',
                                   labels={'Total': 'Valor Total ($)', 'Serviço': 'Valor Serviços ($)'})
                fig_total.update_traces(
                    hovertemplate="<b>%{y}</b><br>Total: $%{x:,.2f}<br>Serviços: $%{marker.color:,.2f}")
                st.plotly_chart(fig_total, use_container_width=True)

            with tab2:
                payment_count = valid_payments['Pagamento'].value_counts().reset_index()
                payment_count.columns = ['Pagamento', 'Qtd Usos']

                # Calcular porcentagem para o gráfico
                total = payment_count['Qtd Usos'].sum()
                payment_count['% Uso'] = (payment_count['Qtd Usos'] / total * 100).round(2)

                fig_qtd = px.bar(payment_count.sort_values('Qtd Usos'),
                                 x='Qtd Usos', y='Pagamento',
                                 title='Quantidade de Usos por Método de Pagamento',
                                 color='Qtd Usos',
                                 color_continuous_scale='Peach',
                                 labels={'Qtd Usos': 'Quantidade de Usos'},
                                 text='% Uso')

                fig_qtd.update_traces(
                    texttemplate='%{text}%',
                    textposition='outside',
                    hovertemplate="<b>%{y}</b><br>Usos: %{x}<br>% do Total: %{text}%"
                )
                st.plotly_chart(fig_qtd, use_container_width=True)

        if not invalid_payments.empty:
            st.warning("Pagamentos inválidos encontrados:")
            st.dataframe(invalid_payments[['Nome', 'Data', 'Cliente', 'Pagamento']])

        st.header("📅 Análise por Dia da Semana")
        day_summary = completed_services.groupby('Dia').agg({
            'Serviço': ['count', 'sum'],
            'Gorjeta': 'sum',
            'Pets': 'sum',
            'Lucro Empresa': 'sum'
        }).reset_index()
        day_summary.columns = ['Dia', 'Atendimentos', 'Total Serviços', 'Total Gorjetas', 'Total Pets', 'Lucro Empresa']
        day_summary['Dia'] = pd.Categorical(day_summary['Dia'],
                                            categories=['Domingo', 'Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta',
                                                        'Sábado'],
                                            ordered=True)
        day_summary = day_summary.sort_values('Dia')

        # Formatar valores monetários para exibição
        day_summary_display = day_summary.copy()
        day_summary_display['Total Serviços'] = day_summary_display['Total Serviços'].apply(format_currency)
        day_summary_display['Total Gorjetas'] = day_summary_display['Total Gorjetas'].apply(format_currency)
        day_summary_display['Lucro Empresa'] = day_summary_display['Lucro Empresa'].apply(format_currency)

        st.dataframe(day_summary_display)

        fig7 = px.bar(day_summary, x='Dia', y='Atendimentos',
                      title='Atendimentos por Dia da Semana',
                      labels={'Atendimentos': 'Quantidade'})
        fig7.update_traces(hovertemplate="<b>%{x}</b><br>Atendimentos: %{y}")
        st.plotly_chart(fig7, use_container_width=True)

        st.header("📤 Exportar Dados")
        col1, col2, col3 = st.columns(3)

        with col1:
            if st.button("Exportar CSV"):
                csv = data.to_csv(index=False).encode('utf-8')
                st.download_button("📁 Baixar CSV", data=csv, file_name="servicos_tecnicos.csv", mime="text/csv")

        with col2:
            if st.button("Exportar Relatório PDF"):
                # Garante que as colunas necessárias existam antes de criar o PDF
                if 'Pagamento Tecnico' not in data.columns:
                    data['Pagamento Tecnico'] = 0
                if 'Lucro Empresa' not in data.columns:
                    data['Lucro Empresa'] = data['Serviço'] + data['Gorjeta'] - data['Pagamento Tecnico']

                pdf = create_pdf(data)
                pdf_bytes = pdf.output(dest='S').encode('latin-1')
                st.download_button(
                    label="📄 Baixar Relatório Completo",
                    data=pdf_bytes,
                    file_name="relatorio_servicos_tecnicos.pdf",
                    mime="application/pdf"
                )

        with col3:
            # Verifica se apenas um técnico e uma semana estão selecionados
            if len(selected_techs) == 1 and len(selected_weeks) == 1:
                tech_name = selected_techs[0]
                week = selected_weeks[0]

                # Filtra os dados para o técnico e semana selecionados
                tech_data = completed_services[
                    (completed_services['Nome'] == tech_name) &
                    (completed_services['Semana'] == week)
                    ]

                if not tech_data.empty:
                    if st.button("Exportar Recibo Técnico"):
                        pdf = create_tech_payment_receipt(tech_data, tech_name, week)
                        pdf_bytes = pdf.output(dest='S').encode('latin-1')
                        st.download_button(
                            label="🧾 Baixar Recibo de Pagamento",
                            data=pdf_bytes,
                            file_name=f"recibo_pagamento_{tech_name}_{week}.pdf",
                            mime="application/pdf"
                        )
                else:
                    st.warning("Nenhum dado encontrado para o técnico selecionado nesta semana.")
            else:
                st.warning("Selecione apenas um técnico e uma semana para gerar o recibo.")

st.markdown("""
    <style>
    .stMetricValue { font-size: 22px; }
    .stDataFrame th, .stDataFrame td { padding: 8px 10px; }
    .css-1aumxhk { background-color: #f9f9f9; padding: 20px; border-radius: 10px; }
    </style>
""", unsafe_allow_html=True)

st.markdown("""
---
<small>Desenvolvido por Alan Salviano | Análise de Planilhas de Serviços Técnicos</small>
""", unsafe_allow_html=True)
